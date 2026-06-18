#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
抓取内蒙古招生计划网页中前四个科类 -> 院校名称链接 -> 专业计划明细，
并导出 Excel。默认不抓取招生章程，也不导出父级院校列表中未涉及的投档/录取占位字段。

默认入口：
https://www.nm.zsks.cn/25gkwb/25zsjh/gkjh_25_3_0628/jh/jhkl.html
"""

from __future__ import annotations

import argparse
import html
import json
import re
import ssl
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlencode, urljoin, urldefrag
from urllib.request import Request, urlopen

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("缺少依赖 openpyxl，请先运行：python3 -m pip install openpyxl") from exc


DEFAULT_URL = "https://www.nm.zsks.cn/25gkwb/25zsjh/gkjh_25_3_0628/jh/jhkl.html"
JSON_URL_RE = re.compile(r"url\s*:\s*['\"](?P<url>[^'\"]+\.json)['\"]", re.I)
TAG_RE = re.compile(r"<[^>]+>")
BR_RE = re.compile(r"<br\s*/?>", re.I)


DETAIL_COLUMNS = [
    ("kl_order", "科类序号"),
    ("kldm", "科类代码"),
    ("klmc", "科类名称"),
    ("pcdm", "批次代码"),
    ("pcmc", "批次名称"),
    ("yxdh", "院校代号"),
    ("yxmc", "院校名称"),
    ("yx_jhs", "院校招生计划数"),
    ("detail_path", "详情 path"),
    ("detail_url", "院校名称链接"),
    ("title", "详情页标题"),
    ("jhzyzdm", "专业组"),
    ("zydh", "专业代号"),
    ("zymc", "专业名称"),
    ("bhzygs", "包含专业"),
    ("jhxzmc", "计划性质"),
    ("jhlbmc", "计划类别"),
    ("xznx", "学制年限"),
    ("xf", "学费"),
    ("wyyzmc", "外语语种"),
    ("kycs", "是否口试"),
    ("kskmyqzw", "选科要求"),
    ("bz", "专业备注"),
    ("bxdd", "办学地点"),
    ("syjh", "招生计划数"),
]

COLLEGE_COLUMNS = [
    ("kl_order", "科类序号"),
    ("kldm", "科类代码"),
    ("klmc", "科类名称"),
    ("pcdm", "批次代码"),
    ("pcmc", "批次名称"),
    ("yxdh", "院校代号"),
    ("yxmc", "院校名称"),
    ("jhs", "院校招生计划数"),
    ("path", "详情 path"),
    ("detail_url", "院校名称链接"),
]

INFO_COLUMNS = [
    ("kldm", "科类代码"),
    ("klmc", "科类名称"),
    ("path", "操作 path"),
    ("pxh", "排序号"),
]

TEXT_FIELDS = {
    "kldm",
    "pcdm",
    "yxdh",
    "path",
    "detail_path",
    "jhzyzdm",
    "zydh",
    "detail_url",
}
NUMBER_FIELDS = {"kl_order", "jhs", "yx_jhs", "syjh"}


def fetch_bytes(url: str, timeout: int) -> tuple[bytes, str | None]:
    request = Request(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0 Safari/537.36"
            ),
            "Accept": "text/html,application/json;q=0.9,*/*;q=0.8",
        },
    )
    context = None
    if url.lower().startswith("https://"):
        context = ssl.create_default_context()
        # 兼容该站点偏旧的 TLS 参数。
        try:
            context.set_ciphers("DEFAULT:@SECLEVEL=1")
        except ssl.SSLError:
            pass

    with urlopen(request, timeout=timeout, context=context) as response:
        charset = response.headers.get_content_charset()
        return response.read(), charset


def decode_text(raw: bytes, charset: str | None = None) -> str:
    for encoding in [charset, "utf-8-sig", "utf-8", "gb18030"]:
        if not encoding:
            continue
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def fetch_text(url: str, timeout: int) -> str:
    raw, charset = fetch_bytes(url, timeout)
    return decode_text(raw, charset)


def load_json_from_url(url: str, timeout: int) -> list[dict[str, Any]]:
    data = json.loads(fetch_text(url, timeout))
    if not isinstance(data, list):
        raise ValueError(f"JSON 数据不是列表：{url}")
    return data


def load_json_from_file(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if not isinstance(data, list):
        raise ValueError(f"JSON 数据不是列表：{path}")
    return data


def find_json_url(html_text: str, page_url: str, expected_name: str) -> str:
    matches = [match.group("url") for match in JSON_URL_RE.finditer(html_text)]
    for match in matches:
        if match.rsplit("/", 1)[-1] == expected_name:
            return urljoin(page_url, match)
    if matches:
        return urljoin(page_url, matches[0])
    raise ValueError(f"没有在页面中找到 JSON 数据地址：{page_url}")


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = html.unescape(text)
    text = BR_RE.sub("\n", text)
    text = TAG_RE.sub("", text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [" ".join(line.split()) for line in text.split("\n")]
    return "\n".join(line for line in lines if line).strip()


def maybe_number(field: str, value: Any) -> Any:
    text = clean_value(value)
    if field not in NUMBER_FIELDS or text == "":
        return text
    try:
        return int(text)
    except ValueError:
        try:
            return float(text)
        except ValueError:
            return text


def make_detail_url(detail_page_url: str, path: Any) -> str:
    return f"{detail_page_url}?{urlencode({'path': clean_value(path)})}"


def select_first_classes(class_rows: list[dict[str, Any]], first_n: int) -> list[dict[str, Any]]:
    if first_n <= 0:
        raise ValueError("--first-n 必须大于 0")
    selected = class_rows[:first_n]
    if not selected:
        raise ValueError("科类数据为空")
    return selected


def build_exports(
    class_rows: list[dict[str, Any]],
    college_rows: list[dict[str, Any]],
    detail_rows: list[dict[str, Any]],
    detail_page_url: str,
    first_n: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    selected_classes = select_first_classes(class_rows, first_n)
    class_by_path = {clean_value(row.get("path")): row for row in selected_classes}
    class_order = {
        clean_value(row.get("path")): idx + 1 for idx, row in enumerate(selected_classes)
    }

    selected_colleges = [
        row for row in college_rows if clean_value(row.get("kldm")) in class_by_path
    ]
    college_by_path = {clean_value(row.get("path")): row for row in selected_colleges}

    details_by_path: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in detail_rows:
        row_path = clean_value(row.get("path"))
        if row_path in college_by_path:
            details_by_path[row_path].append(row)

    college_export: list[dict[str, Any]] = []
    detail_export: list[dict[str, Any]] = []
    unmatched_colleges: list[dict[str, Any]] = []

    for college in selected_colleges:
        college_path = clean_value(college.get("path"))
        class_path = clean_value(college.get("kldm"))
        detail_url = make_detail_url(detail_page_url, college_path)
        college_item = {
            "kl_order": class_order[class_path],
            "kldm": class_path,
            "klmc": clean_value(college.get("klmc")),
            "pcdm": clean_value(college.get("pcdm")),
            "pcmc": clean_value(college.get("pcmc")),
            "yxdh": clean_value(college.get("yxdh")),
            "yxmc": clean_value(college.get("yxmc")),
            "jhs": college.get("jhs"),
            "path": college_path,
            "detail_url": detail_url,
        }
        college_export.append(college_item)

        matched_details = details_by_path.get(college_path, [])
        if not matched_details:
            unmatched_colleges.append(college_item)
            continue

        for detail in matched_details:
            detail_export.append(
                {
                    "kl_order": class_order[class_path],
                    "kldm": class_path,
                    "klmc": clean_value(detail.get("klmc") or college.get("klmc")),
                    "pcdm": clean_value(college.get("pcdm")),
                    "pcmc": clean_value(detail.get("pcmc") or college.get("pcmc")),
                    "yxdh": clean_value(detail.get("yxdh") or college.get("yxdh")),
                    "yxmc": clean_value(detail.get("yxmc") or college.get("yxmc")),
                    "yx_jhs": college.get("jhs"),
                    "detail_path": college_path,
                    "detail_url": detail_url,
                    "title": detail.get("title"),
                    "jhzyzdm": detail.get("jhzyzdm"),
                    "zydh": detail.get("zydh"),
                    "zymc": detail.get("zymc"),
                    "bhzygs": detail.get("bhzygs"),
                    "jhxzmc": detail.get("jhxzmc"),
                    "jhlbmc": detail.get("jhlbmc"),
                    "xznx": detail.get("xznx"),
                    "xf": detail.get("xf"),
                    "wyyzmc": detail.get("wyyzmc"),
                    "kycs": detail.get("kycs"),
                    "kskmyqzw": detail.get("kskmyqzw"),
                    "bz": detail.get("bz"),
                    "bxdd": detail.get("bxdd"),
                    "syjh": detail.get("syjh"),
                }
            )

    selected_class_export = [
        {
            "kldm": clean_value(row.get("kldm")),
            "klmc": clean_value(row.get("klmc")),
            "path": clean_value(row.get("path")),
            "pxh": row.get("pxh"),
        }
        for row in selected_classes
    ]
    return selected_class_export, college_export, detail_export, unmatched_colleges


def append_table(ws, columns: list[tuple[str, str]], rows: list[dict[str, Any]]) -> None:
    ws.append([title for _, title in columns])
    for row in rows:
        ws.append([maybe_number(field, row.get(field)) for field, _ in columns])


def style_sheet(ws, wrap_columns: set[str], columns: list[tuple[str, str]]) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wrap_indexes = {
        idx + 1 for idx, (field, _) in enumerate(columns) if field in wrap_columns
    }
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=cell.column in wrap_indexes,
            )

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        header = ws.cell(row=1, column=col_idx).value
        sample_values = [header] + [cell.value for cell in column_cells[1:250]]
        max_len = max(len(str(value)) for value in sample_values if value is not None)
        width = min(max(max_len + 2, 10), 48)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def set_text_formats(ws, columns: list[tuple[str, str]]) -> None:
    for col_idx, (field, _) in enumerate(columns, start=1):
        if field in TEXT_FIELDS:
            for col_cells in ws.iter_cols(
                min_col=col_idx,
                max_col=col_idx,
                min_row=2,
                max_row=ws.max_row,
            ):
                for cell in col_cells:
                    cell.number_format = "@"


def build_workbook(
    page_url: str,
    class_rows: list[dict[str, Any]],
    college_rows: list[dict[str, Any]],
    detail_rows: list[dict[str, Any]],
    class_json_url: str,
    college_json_url: str,
    detail_json_url: str,
    detail_page_url: str,
    first_n: int,
) -> Workbook:
    selected_classes, colleges, details, unmatched = build_exports(
        class_rows=class_rows,
        college_rows=college_rows,
        detail_rows=detail_rows,
        detail_page_url=detail_page_url,
        first_n=first_n,
    )

    wb = Workbook()
    detail_ws = wb.active
    detail_ws.title = "专业计划明细"
    append_table(detail_ws, DETAIL_COLUMNS, details)
    style_sheet(
        detail_ws,
        wrap_columns={"zymc", "bhzygs", "bz", "bxdd", "title", "detail_url"},
        columns=DETAIL_COLUMNS,
    )
    set_text_formats(detail_ws, DETAIL_COLUMNS)

    college_ws = wb.create_sheet("院校汇总")
    append_table(college_ws, COLLEGE_COLUMNS, colleges)
    style_sheet(college_ws, wrap_columns={"detail_url"}, columns=COLLEGE_COLUMNS)
    set_text_formats(college_ws, COLLEGE_COLUMNS)

    class_ws = wb.create_sheet("科类")
    append_table(class_ws, INFO_COLUMNS, selected_classes)
    style_sheet(class_ws, wrap_columns=set(), columns=INFO_COLUMNS)
    set_text_formats(class_ws, INFO_COLUMNS)

    if unmatched:
        unmatched_ws = wb.create_sheet("未匹配院校")
        append_table(unmatched_ws, COLLEGE_COLUMNS, unmatched)
        style_sheet(unmatched_ws, wrap_columns={"detail_url"}, columns=COLLEGE_COLUMNS)
        set_text_formats(unmatched_ws, COLLEGE_COLUMNS)

    info_ws = wb.create_sheet("说明")
    info_rows = [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("入口页面", page_url),
        ("科类 JSON", class_json_url),
        ("院校 JSON", college_json_url),
        ("专业 JSON", detail_json_url),
        ("选择科类数量", first_n),
        ("选择科类", "、".join(row["klmc"] for row in selected_classes)),
        ("院校链接数量", len(colleges)),
        ("专业计划行数", len(details)),
        ("未匹配院校数量", len(unmatched)),
        ("排除内容", "招生章程 zszc；父级院校列表中未涉及的投档/录取占位字段"),
    ]
    for row in info_rows:
        info_ws.append(row)
    info_ws.column_dimensions["A"].width = 20
    info_ws.column_dimensions["B"].width = 110
    for cell in info_ws["A"]:
        cell.font = Font(bold=True)
    for row in info_ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    return wb


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="抓取前 N 个科类下院校名称链接里的专业计划数据，并导出 .xlsx。"
    )
    parser.add_argument("--url", default=DEFAULT_URL, help="入口科类页 URL")
    parser.add_argument(
        "-o",
        "--output",
        default="nm_zsks_2025_bk_jh_first4.xlsx",
        help="输出 Excel 文件路径",
    )
    parser.add_argument("--first-n", type=int, default=4, help="抓取前几个科类")
    parser.add_argument("--timeout", type=int, default=30, help="网络请求超时时间（秒）")
    parser.add_argument("--class-json", type=Path, help="可选：本地科类 JSON 文件")
    parser.add_argument("--college-json", type=Path, help="可选：本地院校 JSON 文件")
    parser.add_argument("--detail-json", type=Path, help="可选：本地专业计划 JSON 文件")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    page_url = urldefrag(args.url)[0]
    output_path = Path(args.output)

    print(f"入口页面：{page_url}")
    class_html = fetch_text(page_url, args.timeout)
    class_json_url = find_json_url(class_html, page_url, "jhkl.json")

    college_page_url = urljoin(page_url, "jhyx.html")
    college_html = fetch_text(college_page_url, args.timeout)
    college_json_url = find_json_url(college_html, college_page_url, "jhyx.json")

    detail_page_url = urljoin(page_url, "jhzy.html")
    detail_html = fetch_text(detail_page_url, args.timeout)
    detail_json_url = find_json_url(detail_html, detail_page_url, "jhzy.json")

    print(f"科类数据：{class_json_url}")
    print(f"院校数据：{college_json_url}")
    print(f"专业数据：{detail_json_url}")

    class_rows = (
        load_json_from_file(args.class_json)
        if args.class_json
        else load_json_from_url(class_json_url, args.timeout)
    )
    college_rows = (
        load_json_from_file(args.college_json)
        if args.college_json
        else load_json_from_url(college_json_url, args.timeout)
    )
    detail_rows = (
        load_json_from_file(args.detail_json)
        if args.detail_json
        else load_json_from_url(detail_json_url, args.timeout)
    )

    workbook = build_workbook(
        page_url=page_url,
        class_rows=class_rows,
        college_rows=college_rows,
        detail_rows=detail_rows,
        class_json_url=str(args.class_json or class_json_url),
        college_json_url=str(args.college_json or college_json_url),
        detail_json_url=str(args.detail_json or detail_json_url),
        detail_page_url=detail_page_url,
        first_n=args.first_n,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    print(f"完成：{output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
